import logging
import openpyxl
from telegram import Update
from telegram.ext import filters, MessageHandler, ApplicationBuilder, CommandHandler, ContextTypes
from datetime import datetime
from my_token import token

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(
        chat_id=update.effective_chat.id, 
        text="Please talk to me!"
        )

async def help(update: Update, context: ContextTypes.DEFAULT_TYPE):

    await context.bot.send_message(
        chat_id=update.effective_chat.id, 
        text="Syntax:\n/hisp,task name,time spent,status\n/ebrs,task name,time spent,status"
        )

async def list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    
    with open("Daily timesheet.xlsx", "rb") as f:
        await context.bot.send_document(chat_id=update.effective_chat.id, document=f)

async def hisp(update: Update, context: ContextTypes.DEFAULT_TYPE):

    # Open an existing Excel file
    workbook = openpyxl.load_workbook(filename='Daily timesheet.xlsx')
    
    text_msg = update.message.text

    keywords = text_msg.split(",")
    task_name = keywords[1]
    time_spent = keywords[2]
    status = keywords[3]

    current_datetime = datetime.now().strftime("%d/%m/%Y")

    # Select the sheet
    worksheet = workbook['HISP records']

    # get the current last row number
    last_row = worksheet.max_row   
     
     # add data to the worksheet
    worksheet.cell(row=last_row+1, column=1).value = current_datetime
    worksheet.cell(row=last_row+1, column=2).value = task_name
    worksheet.cell(row=last_row+1, column=3).value = time_spent
    worksheet.cell(row=last_row+1, column=4).value = status

    workbook.save(filename='Daily timesheet.xlsx')
   
   
    await context.bot.send_message(
        chat_id=update.effective_chat.id, 
        text="The message has been saved to the HISP records sheet."
        )


async def ebrs(update: Update, context: ContextTypes.DEFAULT_TYPE):

    # Open an existing Excel file
    workbook = openpyxl.load_workbook(filename='Daily timesheet.xlsx')
    
    text_msg = update.message.text

    keywords = text_msg.split(",")
    task_name = keywords[1]
    time_spent = keywords[2]
    status = keywords[3]

    current_datetime = datetime.now().strftime("%d/%m/%Y")

    # Select the sheet
    worksheet = workbook['eBRS records']

    # get the current last row number
    last_row = worksheet.max_row   
     
     # add data to the worksheet
    worksheet.cell(row=last_row+1, column=1).value = current_datetime
    worksheet.cell(row=last_row+1, column=2).value = task_name
    worksheet.cell(row=last_row+1, column=3).value = time_spent
    worksheet.cell(row=last_row+1, column=4).value = status

    workbook.save(filename='Daily timesheet.xlsx')
   
   
    await context.bot.send_message(
        chat_id=update.effective_chat.id, 
        text="The message has been saved to the eBRS records sheet."
        )


if __name__ == '__main__':
    # Jim's token
    application = ApplicationBuilder().token(token.TELEGRAM_BOT_TOKEN).build()
    
    start_handler = CommandHandler('start', start)
    hisp_handler = CommandHandler('hisp', hisp)
    ebrs_handler = CommandHandler('ebrs', ebrs)
    help_handler = CommandHandler('help', help)
    list_handler = CommandHandler('list', list)
    # echo_handler = MessageHandler(filters.TEXT & (~filters.COMMAND), spent)
    
    application.add_handler(start_handler)
    application.add_handler(hisp_handler)
    application.add_handler(ebrs_handler)
    application.add_handler(help_handler)
    application.add_handler(list_handler)
    
    application.run_polling()